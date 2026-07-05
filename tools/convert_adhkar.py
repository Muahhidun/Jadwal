# -*- coding: utf-8 -*-
"""Конвертер: «Jadwal — сбор зикров.xlsx» → app/assets/data/adhkar.json

Запуск:  python3 tools/convert_adhkar.py [--only-approved]
По умолчанию берёт все строки (разработка); с флагом --only-approved — только «проверено» (релиз).
"""
import json, sys, unicodedata
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Jadwal — сбор зикров.xlsx"
OUT = ROOT / "app" / "assets" / "data" / "adhkar.json"

COLLECTIONS = {
    "Утренние зикры": ("morning", "Утренние зикры", "Таңғы зікірлер"),
    "Вечерние зикры": ("evening", "Вечерние зикры", "Кешкі зікірлер"),
    "После намаза": ("after_prayer", "Зикры после намаза", "Намаздан кейінгі зікірлер"),
    "Перед сном": ("before_sleep", "Перед сном", "Ұйықтар алдында"),
    "Пятница": ("friday", "Пятница", "Жұма"),
    "Другое": ("other", "Другое", "Басқа"),
}

def clean(v):
    if v is None:
        return None
    s = unicodedata.normalize("NFC", str(v)).strip()
    return s or None

def main():
    only_approved = "--only-approved" in sys.argv
    ws = load_workbook(XLSX, data_only=True)["Зикры"]
    cols = {}
    skipped = 0
    for r in range(2, ws.max_row + 1):
        ar = clean(ws.cell(r, 3).value)
        if not ar:
            continue
        status = (clean(ws.cell(r, 12).value) or "черновик").lower()
        if only_approved and status != "проверено":
            skipped += 1
            continue
        name = clean(ws.cell(r, 1).value)
        if name not in COLLECTIONS:
            sys.exit(f"Строка {r}: неизвестный сборник «{name}»")
        cid, title_ru, title_kz = COLLECTIONS[name]
        col = cols.setdefault(cid, {"id": cid, "titleRu": title_ru, "titleKz": title_kz, "items": []})
        source = clean(ws.cell(r, 9).value)
        if not source:
            sys.exit(f"Строка {r}: у зикра нет источника — это нарушает правила проекта")
        col["items"].append({
            "order": int(ws.cell(r, 2).value or 0),
            "ar": ar,
            "translit": clean(ws.cell(r, 4).value),
            "ru": clean(ws.cell(r, 5).value),
            "kz": clean(ws.cell(r, 6).value),
            "fazRu": clean(ws.cell(r, 7).value),
            "fazKz": clean(ws.cell(r, 8).value),
            "source": source,
            "repeat": int(ws.cell(r, 10).value or 1),
            "note": clean(ws.cell(r, 11).value),
            "status": status,
        })
    for col in cols.values():
        col["items"].sort(key=lambda i: i["order"])
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"collections": list(cols.values())}, ensure_ascii=False, indent=1), encoding="utf-8")
    total = sum(len(c["items"]) for c in cols.values())
    print(f"OK: {total} зикров в {len(cols)} сборниках → {OUT.relative_to(ROOT)}" + (f" (пропущено не-проверенных: {skipped})" if skipped else ""))

if __name__ == "__main__":
    main()
